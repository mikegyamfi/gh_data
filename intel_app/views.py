import hashlib
import hmac
import json
from datetime import datetime

import pandas as pd
from decouple import config
from django.contrib.auth.forms import PasswordResetForm
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
import requests
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt

from . import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from . import helper, models
from .forms import UploadFileForm
from .models import CustomUser


# Create your views here.
def home(request):
    if request.user.is_authenticated:
        user = models.CustomUser.objects.get(id=request.user.id)
        wallet_balance = user.wallet
        context = {
            "wallet_balance": wallet_balance,
        }
        return render(request, "layouts/index.html", context=context)
    return render(request, "layouts/index.html")


@login_required(login_url='login')
def agent_upgrade(request):
    user = request.user
    fee = models.AdminInfo.objects.first().agent_registration_fee
    reference = helper.ref_generator()

    models.Payment.objects.create(
        user=user,
        reference=reference,
        amount=fee,
        transaction_date=datetime.now(),
        transaction_status="Pending",
    )

    return render(request, "layouts/services/agent_upgrade.html", {
        "fee": fee,
        "reference": reference,
        "public_key": config("PAYSTACK_PUBLIC_KEY"),
        "amount_pesewas": int(fee * 100),
    })


def services(request):
    return render(request, "layouts/services.html")


@login_required(login_url='login')
def pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        print(phone_number)
        print(amount)
        print(reference)
        bundle = models.IshareBundlePrice.objects.get(
            price=float(amount)).bundle_volume if user.status == "User" else models.AgentIshareBundlePrice.objects.get(
            price=float(amount)).bundle_volume
        print(bundle)
        send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, reference)
        data = send_bundle_response.json()
        print(data)

        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        if send_bundle_response.status_code == 200:
            if data["code"] == "0000":
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Completed"
                )
                new_transaction.save()
                user.wallet -= float(amount)
                user.wallet = round(user.wallet, 2)
                user.save()
                receiver_message = f"{bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                sms_message = f"{bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using GH Data.\n\nGH Data"

                num_without_0 = phone_number[1:]
                print(num_without_0)
                receiver_body = {
                    'recipient': f"233{num_without_0}",
                    'sender_id': 'GH DATA',
                    'message': receiver_message
                }

                response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                print(response.text)

                sms_body = {
                    'recipient': f"233{request.user.phone}",
                    'sender_id': 'GH DATA',
                    'message': sms_message
                }

                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                print(response.text)

                return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
            else:
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Failed"
                )
                new_transaction.save()
                return JsonResponse({'status': 'Something went wrong'})
    return redirect('airtel-tigo')


@login_required(login_url='login')
def airtel_tigo(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.IShareBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        form = forms.IShareBundleForm(data=request.POST, status=status)
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        print("payment saved")
        print("form valid")
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        print(offer)
        bundle = models.IshareBundlePrice.objects.get(
            price=float(offer)).bundle_volume if user.status == "User" else models.AgentIshareBundlePrice.objects.get(
            price=float(offer)).bundle_volume
        new_transaction = models.IShareBundleTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
            transaction_status="Pending"
        )
        print("created")
        new_transaction.save()

        print("===========================")
        print(phone_number)
        print(bundle)
        send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
        data = send_bundle_response.json()

        print(data)

        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'

        if send_bundle_response.status_code == 200:
            if data["code"] == "0000":
                transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
                print("got here")
                print(transaction_to_be_updated.transaction_status)
                transaction_to_be_updated.transaction_status = "Completed"
                transaction_to_be_updated.save()
                print(request.user.phone)
                print("***********")
                receiver_message = f"{bundle}MB has been credited to you by {request.user.phone}.\nReference: {payment_reference}\n"
                sms_message = f"{bundle}MB has been credited to {phone_number}.\nReference: {payment_reference}\nThank you for using GH Data.\n\nGH Data"

                num_without_0 = phone_number[1:]
                print(num_without_0)
                receiver_body = {
                    'recipient': f"233{num_without_0}",
                    'sender_id': 'GH DATA',
                    'message': receiver_message
                }

                response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                print(response.text)

                sms_body = {
                    'recipient': f"233{request.user.phone}",
                    'sender_id': 'GH DATA',
                    'message': sms_message
                }

                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                print(response.text)

                return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
            else:
                transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
                transaction_to_be_updated.transaction_status = "Failed"
                new_transaction.save()
                sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using GH Data.\n\nGH Data"

                sms_body = {
                    'recipient': f"233{request.user.phone}",
                    'sender_id': 'GH DATA',
                    'message': sms_message
                }
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
                # r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={phone_number}&from=GH DATA GH&sms={receiver_message}"
                # response = requests.request("GET", url=r_sms_url)
                # print(response.text)
                return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
        else:
            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
            transaction_to_be_updated.transaction_status = "Failed"
            new_transaction.save()
            sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using GH Data.\n\nGH Data"

            sms_body = {
                'recipient': f'233{request.user.phone}',
                'sender_id': 'GH DATA',
                'message': sms_message
            }

            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

            print(response.text)
            return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/at.html", context=context)


@login_required(login_url='login')
def mtn_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        admin = models.AdminInfo.objects.filter().first().phone_number

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        bundle = models.MTNBundlePrice.objects.get(
            price=float(amount)).bundle_volume if user.status == "User" else models.AgentMTNBundlePrice.objects.get(
            price=float(amount)).bundle_volume
        print(bundle)
        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.wallet = round(user.wallet, 2)
        user.save()
        admin = models.AdminInfo.objects.filter().first().phone_number
        sms_body = {
            'recipient': f"233{admin}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('mtn')


@login_required(login_url='login')
def mtn(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.MTNForm(status=status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Completed"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")

        bundle = models.MTNBundlePrice.objects.get(
            price=float(offer)).bundle_volume if user.status == "User" else models.AgentMTNBundlePrice.objects.get(
            price=float(offer)).bundle_volume

        print(phone_number)
        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
        admin = models.AdminInfo.objects.filter().first().phone_number
        sms_body = {
            'recipient': f"233{admin}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {'form': form, "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/mtn.html", context=context)


@login_required(login_url='login')
def big_time_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        bundle = models.BigTimeBundlePrice.objects.get(
            price=float(amount)).bundle_volume if user.status == "User" else models.AgentBigTimeBundlePrice.objects.get(
            price=float(amount)).bundle_volume
        print(bundle)
        print(bundle)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.wallet = round(user.wallet, 2)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('big_time')


@login_required(login_url='login')
def big_time(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.BigTimeBundleForm(status)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email

    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Pending"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        bundle = models.BigTimeBundlePrice.objects.get(
            price=float(offer)).bundle_volume if user.status == "User" else models.AgentBigTimeBundlePrice.objects.get(
            price=float(offer)).bundle_volume
        print(bundle)
        print(phone_number)
        new_mtn_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, 'id': db_user_id,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/big_time.html", context=context)


@login_required(login_url='login')
def history(request):
    user_transactions = models.IShareBundleTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "AirtelTigo Transactions"
    net = "tigo"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def mtn_history(request):
    user_transactions = models.MTNTransaction.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "MTN Transactions"
    net = "mtn"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def big_time_history(request):
    user_transactions = models.BigTimeTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Big Time Transactions"
    net = "bt"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def admin_bt_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.BigTimeTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/bt_admin.html", context=context)


@login_required(login_url='login')
def bt_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.BigTimeTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT BIG TIME transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('bt_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('bt_admin')


def verify_transaction(request, reference):
    if request.method == "GET":
        response = helper.verify_paystack_transaction(reference)
        data = response.json()
        try:
            status = data["data"]["status"]
            amount = data["data"]["amount"]
            api_reference = data["data"]["reference"]
            date = data["data"]["paid_at"]
            real_amount = float(amount) / 100
            print(status)
            print(real_amount)
            print(api_reference)
            print(reference)
            print(date)
        except:
            status = data["status"]
        return JsonResponse({'status': status})


from django.db.models import FloatField
from django.db.models.functions import Cast, Substr, Length


def change_excel_status(request, status, to_change_to):
    if request.user.is_superuser and request.user.is_staff:
        transactions = models.MTNTransaction.objects.filter(
            transaction_status=status) if to_change_to != "Completed" else models.MTNTransaction.objects.filter(
            transaction_status=status).order_by('transaction_date')[:10]
        for transaction in transactions:
            transaction.transaction_status = to_change_to
            transaction.save()
            if to_change_to == "Completed":
                transaction_number = transaction.user.phone
                sms_headers = {
                    'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                    'Content-Type': 'application/json'
                }

                sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                sms_message = f"{transaction.bundle_number} has been credited with {transaction.offer}.\nTransaction Reference: {transaction.reference}"

                sms_body = {
                    'recipient': f"233{transaction_number}",
                    'sender_id': 'GH DATA',
                    'message': sms_message
                }
                try:
                    response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                    print(response.text)
                except:
                    messages.success(request, f"Transaction Completed")
                    return redirect('mtn_admin', status=status)
            else:
                messages.success(request, f"Status changed from {status} to {to_change_to}")
                return redirect("mtn_admin", status=status)
        messages.success(request, f"Status changed from {status} to {to_change_to}")
        return redirect("mtn_admin", status=status)
    else:
        messages.info(request, "Link Broken")
        return redirect("home")


from django.db.models import FloatField
from django.db.models.functions import Cast, Length, Substr


@login_required(login_url='login')
def admin_mtn_history(request, status):
    if request.user.is_staff and request.user.is_superuser:
        if request.method == "POST":
            from io import BytesIO
            from openpyxl import load_workbook
            from django.http import HttpResponse
            import datetime

            # Get the uploaded Excel file
            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                messages.error(request, "No Excel file found")
                return redirect('mtn_admin', status=status)

            # Load the uploaded Excel file into memory
            excel_buffer = BytesIO(uploaded_file.read())
            book = load_workbook(excel_buffer)
            sheet = book.active  # Assuming data is on the active sheet

            # Column indices for "RECIPIENT" and "DATA"
            recipient_col_index = 1  # Update if necessary
            data_col_index = 2  # Update if necessary

            # Query transactions with the current status
            queryset = models.MTNTransaction.objects.filter(transaction_status=status).annotate(
                offer_value=Cast(Substr('offer', 1, Length('offer') - 2), FloatField())
            ).order_by('-offer_value')

            # Start from row 2 (assuming headers in row 1)
            start_row = 2

            for record in queryset:
                recipient_value = f"0{record.bundle_number}"
                data_value = record.offer
                cleaned_data_value = float(data_value.replace('MB', ''))
                data_value_gb = round(cleaned_data_value / 1000, 2)

                # Find next empty row
                while sheet.cell(row=start_row, column=recipient_col_index).value is not None:
                    start_row += 1

                # Write data to cells
                sheet.cell(row=start_row, column=recipient_col_index, value=recipient_value)
                sheet.cell(row=start_row, column=data_col_index, value=data_value_gb)

                # Update the record status if necessary
                if status == "Pending":
                    record.transaction_status = "Processing"
                    record.save()
                # For "Processing" status, we do not change the status
                # You can add more logic here if needed

            # Save and return the modified Excel file
            excel_buffer.seek(0)
            book.save(excel_buffer)
            excel_buffer.seek(0)
            response = HttpResponse(excel_buffer.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response[
                'Content-Disposition'] = f'attachment; filename={datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S")}.xlsx'
            return response

        # For GET requests, display transactions based on status
        all_txns = models.MTNTransaction.objects.filter(transaction_status=status).order_by('-transaction_date')[:800]
        context = {'txns': all_txns, 'status': status}
        return render(request, "layouts/services/mtn_admin.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.MTNTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your account has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.bundle_number}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        return redirect('mtn_admin', status="Pending")


def credit_user(request):
    if request.user.is_superuser:
        form = forms.CreditUserForm()
        if request.method == "POST":
            form = forms.CreditUserForm(request.POST)
            if form.is_valid():
                user = form.cleaned_data["user"]
                amount = form.cleaned_data["amount"]
                print(user)
                print(amount)
                user_needed = models.CustomUser.objects.get(username=user)
                if user_needed.wallet is None:
                    user_needed.wallet = float(amount)
                else:
                    user_needed.wallet += float(amount)
                user_needed.save()
                print(user_needed.username)
                messages.success(request, "Crediting Successful")
                return redirect('credit_user')
        context = {'form': form}
        return render(request, "layouts/services/credit.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def topup_info(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        amount = request.POST.get("amount")
        print(amount)
        reference = helper.top_up_ref_generator()
        new_topup_request = models.TopUpRequest.objects.create(
            user=request.user,
            amount=amount,
            reference=reference,
        )
        new_topup_request.save()

        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"A top up request has been placed.\nGHS{amount} for {user}.\nReference: {reference}"

        sms_body = {
            'recipient': f"233{admin}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            pass
        messages.success(request,
                         f"Your Request has been sent successfully. Kindly go on to pay to {admin} and use the reference stated as reference. Reference: {reference}")
        return redirect("request_successful", reference)
    db_user_id = request.user.id
    user_email = request.user.email
    reference = helper.ref_generator()
    context = {'id': db_user_id, "ref": reference, "email": user_email}
    return render(request, "layouts/topup-info.html", context=context)


@login_required(login_url='login')
def request_successful(request, reference):
    admin = models.AdminInfo.objects.filter().first()
    context = {
        "name": admin.name,
        "number": f"0{admin.momo_number}",
        "channel": admin.payment_channel,
        "reference": reference
    }
    return render(request, "layouts/services/request_successful.html", context=context)


def topup_list(request):
    if request.user.is_superuser:
        topup_requests = models.TopUpRequest.objects.all().order_by('date').reverse()
        context = {
            'requests': topup_requests,
        }
        return render(request, "layouts/services/topup_list.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def credit_user_from_list(request, reference):
    if request.user.is_superuser:
        crediting = models.TopUpRequest.objects.filter(reference=reference).first()
        user = crediting.user
        custom_user = models.CustomUser.objects.get(username=user.username)
        if crediting.status:
            return redirect('topup_list')
        amount = crediting.amount
        print(user)
        print(user.phone)
        print(amount)
        custom_user.wallet += amount
        custom_user.save()
        crediting.status = True
        crediting.credited_at = datetime.now()
        crediting.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nYour wallet has been topped up with GHS{amount}.\nReference: {reference}.\nThank you"

        sms_body = {
            'recipient': f"233{custom_user.phone}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            pass
        messages.success(request, f"{user} has been credited with {amount}")
        return redirect('topup_list')


@login_required(login_url='login')
def afa_registration(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    reference = helper.ref_generator()
    price = models.AdminInfo.objects.filter().first().afa_price
    user_email = request.user.email
    print(price)
    if request.method == "POST":
        form = forms.AFARegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Registration will be done shortly")
    form = forms.AFARegistrationForm()
    context = {'form': form, 'ref': reference, 'price': price, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/afa.html", context=context)


def afa_registration_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        name = request.POST.get("name")
        card_number = request.POST.get("card")
        occupation = request.POST.get("occupation")
        date_of_birth = request.POST.get("birth")
        region = request.POST.get("region")
        location = request.POST.get("location")
        print(region)
        print(location)
        price = models.AdminInfo.objects.filter().first().afa_price

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        new_registration = models.AFARegistration.objects.create(
            user=user,
            reference=reference,
            name=name,
            phone_number=phone_number,
            gh_card_number=card_number,
            occupation=occupation,
            date_of_birth=date_of_birth,
            region=region,
            location=location
        )
        new_registration.save()
        user.wallet -= float(price)
        user.wallet = round(user.wallet, 2)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('home')


def afa_history(request):
    user_transactions = models.AFARegistration.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "AFA Registrations"
    net = "afa"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/afa_history.html", context=context)


@login_required(login_url='login')
def admin_afa_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.AFARegistration.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/afa_admin.html", context=context)


@login_required(login_url='login')
def afa_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.AFARegistration.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"AFA Registration.{txn.phone_number} has been registered.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        messages.success(request, f"Transaction Completed")
        return redirect('afa_admin')


@login_required(login_url='login')
def voda_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.VodafoneTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"{txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
        print(response.text)
        messages.success(request, f"Transaction Completed")
        return redirect('voda_admin')


@login_required(login_url='login')
def voda(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.VodaBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email

    if request.method == "POST":
        payment_reference = request.POST.get("reference")
        amount_paid = request.POST.get("amount")
        new_payment = models.Payment.objects.create(
            user=request.user,
            reference=payment_reference,
            amount=amount_paid,
            transaction_date=datetime.now(),
            transaction_status="Pending"
        )
        new_payment.save()
        phone_number = request.POST.get("phone")
        offer = request.POST.get("amount")
        bundle = models.VodaBundlePrice.objects.get(
            price=float(offer)).bundle_volume if user.status == "User" else models.AgentVodaBundlePrice.objects.get(
            price=float(offer)).bundle_volume

        print(phone_number)
        new_mtn_transaction = models.VodafoneTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=payment_reference,
        )
        new_mtn_transaction.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/voda.html", context=context)


@login_required(login_url='login')
def voda_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        bundle = models.VodaBundlePrice.objects.get(
            price=float(amount)).bundle_volume if user.status == "User" else models.AgentVodaBundlePrice.objects.get(
            price=float(amount)).bundle_volume

        print(bundle)
        new_mtn_transaction = models.VodafoneTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.wallet = round(user.wallet, 2)
        user.save()
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('voda')


@login_required(login_url='login')
def voda_history(request):
    user_transactions = models.VodafoneTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Vodafone Transactions"
    net = "voda"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def admin_voda_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.VodafoneTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/voda_admin.html", context=context)


@csrf_exempt
def paystack_webhook(request):
    if request.method == "POST":
        paystack_secret_key = config("PAYSTACK_SECRET_KEY")
        # print(paystack_secret_key)
        payload = json.loads(request.body)

        paystack_signature = request.headers.get("X-Paystack-Signature")

        if not paystack_secret_key or not paystack_signature:
            return HttpResponse(status=400)

        computed_signature = hmac.new(
            paystack_secret_key.encode('utf-8'),
            request.body,
            hashlib.sha512
        ).hexdigest()

        if computed_signature == paystack_signature:
            print("yes")
            print(payload.get('data'))
            r_data = payload.get('data')
            print(r_data.get('metadata'))
            print(payload.get('event'))
            if payload.get('event') == 'charge.success':
                metadata = r_data.get('metadata')
                receiver = metadata.get('receiver')
                db_id = metadata.get('db_id')
                referer = metadata.get('referrer')
                print(db_id)
                # offer = metadata.get('offer')
                user = models.CustomUser.objects.get(id=int(db_id))
                print(user)
                channel = metadata.get('channel')
                real_amount = metadata.get('real_amount')
                print(real_amount)
                paid_amount = r_data.get('amount')
                slashed_amount = float(paid_amount) / 100
                slashed_amount = (float(paid_amount) / 100) * (1 - 0.0195)
                reference = r_data.get('reference')

                rounded_real_amount = round(float(real_amount))
                rounded_paid_amount = round(float(slashed_amount))

                print(f"reeeeeeeaaaaaaaaal amount: {rounded_real_amount}")
                print(f"paaaaaaaaaaaaaiiddd amount: {rounded_paid_amount}")

                is_within_range = (rounded_real_amount - 5) <= rounded_paid_amount <= (rounded_real_amount + 5)

                if not is_within_range:
                    sms_headers = {
                        'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    sms_message = f"Malicious attempt on webhook. Real amount: {rounded_real_amount} | Paid amount: {rounded_paid_amount}. Referrer: {reference}"

                    sms_body = {
                        'recipient': "233242442147",
                        'sender_id': 'GH DATA',
                        'message': sms_message
                    }
                    try:
                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        print(response.text)
                    except:
                        pass

                    print("not within range")
                    return HttpResponse(200)

                if channel == "topup":
                    try:
                        topup_amount = metadata.get('real_amount')

                        if models.TopUpRequest.objects.filter(user=user, reference=reference).exists():
                            return HttpResponse(status=200)

                        new_payment = models.Payment.objects.create(
                            user=user,
                            reference=reference,
                            amount=paid_amount,
                            transaction_date=datetime.now(),
                            transaction_status="Completed"
                        )
                        new_payment.save()
                        print(user.wallet)
                        user.wallet += float(topup_amount)
                        user.save()
                        print(user.wallet)

                        if models.TopUpRequest.objects.filter(user=user, reference=reference, status=True).exists():
                            return HttpResponse(status=200)

                        new_topup = models.TopUpRequest.objects.create(
                            user=user,
                            reference=reference,
                            amount=topup_amount,
                            status=True,
                        )
                        new_topup.save()

                        sms_headers = {
                            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                            'Content-Type': 'application/json'
                        }

                        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                        sms_message = f"Your GH Data wallet has been credited with GHS{topup_amount}.\nReference: {reference}\n"

                        sms_body = {
                            'recipient': f"233{user.phone}",
                            'sender_id': 'GH DATA',
                            'message': sms_message
                        }
                        try:
                            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                            print(response.text)
                            return HttpResponse(status=200)
                        except:
                            return HttpResponse(status=200)
                    except:
                        return HttpResponse(status=200)
                if channel == "agent":
                    try:
                        user = CustomUser.objects.get(id=int(metadata.get('db_id')))
                        user.status = "Agent"
                        user.save()

                        # mark the Payment record completed
                        payment = models.Payment.objects.filter(
                            reference=reference, user=user
                        ).first()
                        if payment:
                            payment.transaction_status = "Completed"
                            payment.save()

                        # notify via SMS
                        sms_headers = {
                            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                            'Content-Type': 'application/json'
                        }
                        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                        sms_body = {
                            'recipient': f"233{user.phone}",
                            'sender_id': 'GH DATA',
                            'message': f"Congratulations {user.first_name}! Your account has been upgraded to Agent."
                        }
                        try:
                            requests.post(url=sms_url, params=sms_body, headers=sms_headers)
                        except Exception as e:
                            print(e)
                            pass
                    except Exception as e:
                        print(e)
                        pass
                    return HttpResponse(status=200)
                else:
                    return HttpResponse(status=200)
            else:
                return HttpResponse(status=200)
        else:
            return HttpResponse(status=401)
    else:
        return HttpResponse(status=200)


def query_txn(request):
    if request.method == "POST":
        reference = request.POST.get('reference')
        print(reference)

        headers = {
            "api-key": config("API_KEY"),
            "api-secret": config("API_SECRET"),
        }
        response = requests.post(
            url=f"https://console.bestpaygh.com/api/flexi/v1/transaction_detail/{reference.strip()}/", headers=headers)
        data = response.json()
        print(data)
        try:
            print(data["message"])
            messages.info(request, data["message"])
        except:
            print(data['api_response']['message'])
            messages.info(request, data['api_response']['message'])

        return redirect('query_txn')
    return render(request, "layouts/query_txn.html")


from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            user = models.CustomUser.objects.filter(email=data).first()
            current_user = user
            if user:
                subject = "Password Reset Requested"
                email_template_name = "password/password_reset_message.txt"
                c = {
                    "name": user.first_name,
                    "email": user.email,
                    'domain': 'www.ghdata.store',
                    'site_name': 'GH DATA',
                    "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                    "user": user,
                    'token': default_token_generator.make_token(user),
                    'protocol': 'https',
                }
                email = render_to_string(email_template_name, c)

                sms_headers = {
                    'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
                    'Content-Type': 'application/json'
                }

                sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                sms_body = {
                    'recipient': f"233{user.phone}",
                    'sender_id': 'GH DATA',
                    'message': email
                }
                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                print(response.text)
                # requests.get(
                #     f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UnBzemdvanJyUGxhTlJzaVVQaHk&to=0{current_user.phone}&from=GEO_AT&sms={email}")

                return redirect("/password_reset/done/")
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password/password_reset.html",
                  context={"password_reset_form": password_reset_form})


def populate_custom_users_from_excel(request):
    # Read the Excel file using pandas
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']

            # Process the uploaded Excel file
            df = pd.read_excel(excel_file)
            counter = 0
            # Iterate through rows to create CustomUser instances
            for index, row in df.iterrows():
                print(counter)
                # Create a CustomUser instance for each row
                custom_user = models.CustomUser.objects.create(
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    username=str(row['username']),
                    email=row['email'],
                    phone=row['phone'],
                    wallet=float(row['wallet']),
                    status=str(row['status']),
                    password1=row['password1'],
                    password2=row['password2'],
                    is_superuser=row['is_superuser'],
                    is_staff=row['is_staff'],
                    is_active=row['is_active'],
                    password=row['password']
                )

                custom_user.save()

                # group_names = row['groups'].split(',')  # Assuming groups are comma-separated
                # groups = Group.objects.filter(name__in=group_names)
                # custom_user.groups.set(groups)
                #
                # if row['user_permissions']:
                #     permission_ids = [int(pid) for pid in row['user_permissions'].split(',')]
                #     permissions = Permission.objects.filter(id__in=permission_ids)
                #     custom_user.user_permissions.set(permissions)
                print("killed")
                counter = counter + 1
            messages.success(request, 'All done')
    else:
        form = UploadFileForm()
    return render(request, 'layouts/import_users.html', {'form': form})


def delete_custom_users(request):
    models.CustomUser.objects.all().delete()
    return HttpResponseRedirect('Done')
